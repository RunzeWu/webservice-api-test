# ！/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     :2018/12/29 22:00
# @Author   :Yosef
# E-mail    :wurz529@foxmail.com
# File      :do_testcase_excel.py
# Software  :PyCharm Community Edition
import openpyxl
from common.config import ReadConfig
from common import contants
from common import mylog

logger = mylog.get_logger("get test data")


class DoExcel:
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.conf = ReadConfig().get_value("test_case_id", "button")

    def get_init_data(self):
        wb = openpyxl.load_workbook(self.filepath)
        sh = wb["init"]
        mobile = sh.cell(1, 2).value
        return mobile

    def update_init_data(self,value):
        wb = openpyxl.load_workbook(self.filepath)
        sh = wb["init"]
        sh.cell(1, 2).value = value
        wb.save(self.filepath)

    def read_data(self):
        """
        每次都是根据数据库member表号码最大值为参照来替换注册号码数据
        :return:
        """
        wb = openpyxl.load_workbook(self.filepath)
        sh = wb[self.sheet_name]
        logger.info("开始读取{}中的{}数据".format(self.filepath, self.sheet_name))

        col_max = sh.max_column

        testdata_key = []
        for i in range(1, col_max + 1):
            testdata_key.append(sh.cell(1, i).value)

        if self.conf == "all":
            testdatas = []
            row_max = sh.max_row
            for i in range(2, row_max + 1):
                testdata = {}
                for j in range(1, col_max + 1):
                    testdata[testdata_key[j - 1]] = sh.cell(i, j).value
                testdatas.append(testdata)
            logger.info("读取完毕")
        else:
            try:
                conf = eval(self.conf)
            except Exception:
                logger.info("请检查配置文件中[test_case_id]的button的value是否符合要求！！")
                raise Exception("请检查配置文件中[test_case_id]的button的value是否符合要求！！")
            new_conf = []
            for run_case_id in conf:
                run_case_row = run_case_id + 1
                new_conf.append(run_case_row)
            testdatas = []
            for i in new_conf:
                testdata = {}
                for j in range(1, col_max + 1):
                    testdata[testdata_key[j - 1]] = sh.cell(i, j).value
                testdatas.append(testdata)
            logger.info("读取完毕")

        return testdatas

    def write_data(self, row, col, value):
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb[self.sheet_name]

        ws.cell(row, col).value = value
        wb.save(self.filepath)


if __name__ == '__main__':
    excel = DoExcel(contants.case_file, "register").read_data()
    for i in excel:
        print(i, end="\n")
    # DoExcel(ProjectPath().test_data_data(), "register").update_init_data("17751811111")